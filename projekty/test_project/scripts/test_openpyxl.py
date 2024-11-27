# scripts/test_openpyxl.py

from openpyxl import load_workbook

wb = load_workbook('scripts/vek.xlsx')

print(wb.sheetnames)

sheet = wb['List1']

print(sheet['A1'].value)

for row in sheet:
    for cell in row:
        print(cell.value)

